#this project needs to be restructured to be more modular
from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('How can we help today?')
#root.geometry ("900x900")

#c = Checkbutton(root, text="Check this")

wb = load_workbook('MDCheck.xlsx') # work book to be opened. In the furture may need to call to multiple workbooks
Record = open('Patient.txt', 'a')  # txt file to append the records

Pre_History= []    # values from history to be appended to this list
Pre_Examination = []
Pre_Investigations = []
Pre_Management = []
History = ['On review:']
HistoryElements_Vs_Column_A = []
DDXelements_Vs_Column_C = []
ExaminationElements_Vs_Column_E = []
IxElements_Vs_Column_G =[]
MxElements_Vs_Column_I= []
Pre_DDX = []
Differentials = ['History suggest:']
Making_differentials =[]

Sheets = wb.sheetnames # geting sheets in the MDCheck.xlsx

#print(Sheets)


my_listbox = Listbox(root, width=40, selectmode=MULTIPLE, font = ("Helvetica", 12))
my_listbox.grid(column=1, row=1)
#my_listbox.pack()





History_listbox = Listbox(root, width=40, selectmode=MULTIPLE , font = ("Helvetica", 12))
History_listbox.grid(column=2, row=1)
#History_listbox.pack() # this will be a listbox to which history taking items will be appended once reason for are selected

Possible_differentials = Listbox(root, width=40, selectmode=MULTIPLE, font = ("Helvetica", 12))
Possible_differentials.grid(column=3, row=1)
#Possible_differentials.pack()



Examination_listbox = Listbox(root, width=40, selectmode=MULTIPLE, font = ("Helvetica", 12))
Examination_listbox.grid(column=4, row=1)
#Examination_listbox.pack()



Investigations_listbox = Listbox(root, width=40, selectmode=MULTIPLE, font = ("Helvetica", 12))
Investigations_listbox.grid(column=1, row=2)
#Investigations_listbox.pack()



Management_listbox = Listbox(root, width= 50, selectmode=MULTIPLE, font = ("Helvetica", 12))
Management_listbox.grid(column=2, row=2)
#Management_listbox.pack()


List_of_Differentials = Listbox(root, width=40, selectmode=MULTIPLE , font = ("Helvetica", 12))
List_of_Differentials.grid(column=1, row=3)


Consult_Record = Text(root, width = 30 , height= 10, font = ("Helvetica", 12))
Consult_Record.grid(column=2, row=3)
Consult_Record.config(state='normal')


for i in Sheets:
        my_listbox.insert(END, i)

def logger(): # when selections have been made for reasons of attending - the function will lead to population of respective 
        Consult_Record.insert(END,'Review for: '+ '\n')
        for i in my_listbox.curselection():
                History.append(my_listbox.get(i))
                Consult_Record.insert( END, my_listbox.get(i) + '\n')
        for i in History:
                wb = load_workbook('MDCheck.xlsx')
                for sheet in wb:
                        if sheet.title == i:
                                for cells in sheet['A']:
                                        if cells.value != None:
                                                Pre_History.append(cells.value)

                                for cells in sheet['C']:
                                        if cells.value != None:
                                                Pre_DDX.append(cells.value)

                                for cells in sheet['E']:
                                        if cells.value != None:
                                                Pre_Examination.append(cells.value)

                                for cells in sheet['G']:
                                        if cells.value != None:
                                                Pre_Investigations.append(cells.value)
                                for cells in sheet['I']:
                                        if cells.value != None:
                                                Pre_Management.append(cells.value)




                for i in Pre_History:
                        History_listbox.insert(END, i)

                for i in Pre_DDX:
                        Possible_differentials.insert(END, i)
        
                for i in Pre_Examination:
                        Examination_listbox.insert(END, i)


                for i in Pre_Investigations:
                        Investigations_listbox.insert(END, i)
                                

                for i in Pre_Management:
                        Management_listbox.insert(END, i)

        print(History)


def On_review():
        for i in History_listbox.curselection():
                HistoryElements_Vs_Column_A.append(History_listbox.get(i))
                Consult_Record.insert(END, History_listbox.get(i) + '\n')
        
        for i in History:
                #wb = load_workbook('MDCheck.xlsx')
                for sheet in wb:
                        if sheet.title == i:
                                for r in range(1, 500):
                                        if sheet.cell(row=r, column=1).value != None and sheet.cell(row=r, column=2).value != None:
                                                for k in HistoryElements_Vs_Column_A:
                                                        if k == sheet.cell(row=r, column=1).value:
                                                                List_of_Differentials.insert(END, 'History suggest: '+ sheet.cell(row=r, column=1).value)
                                                                Making_differentials = (sheet.cell(row=r, column=2).value.split('|'))
                                                                for i in Making_differentials:
                                                                  List_of_Differentials.insert(END, i)


                                                                



def On_PE():
        for i in Examination_listbox.curselection():
                ExaminationElements_Vs_Column_E.append(Examination_listbox.get(i))
                Consult_Record.insert(END, Examination_listbox.get(i) + '\n')

        
        for i in History:
                #wb = load_workbook('MDCheck.xlsx')
                for sheet in wb:
                        if sheet.title == i:
                                for r in range(1, 500):
                                        if sheet.cell(row=r, column=5).value != None and sheet.cell(row=r, column=6).value != None:
                                                for k in ExaminationElements_Vs_Column_E:
                                                        if k == sheet.cell(row=r, column=5).value:
                                                                List_of_Differentials.insert(END, 'History suggest: '+ sheet.cell(row=r, column=5).value)
                                                                Making_differentials = (sheet.cell(row=r, column=6).value.split('|'))
                                                                for i in Making_differentials:
                                                                  List_of_Differentials.insert(END, i)

def DDX():
        for i in Possible_differentials.curselection():
                DDXelements_Vs_Column_C.append(Possible_differentials.get(i))
                Consult_Record.insert(END, Possible_differentials.get(i) + '\n')

        
        for i in History:
                #wb = load_workbook('MDCheck.xlsx')
                for sheet in wb:
                        if sheet.title == i:
                                for r in range(1, 500):
                                        if sheet.cell(row=r, column=3).value != None and sheet.cell(row=r, column=4).value != None:
                                                for k in DDXelements_Vs_Column_C:
                                                        if k == sheet.cell(row=r, column=3).value:
                                                                List_of_Differentials.insert(END, 'History suggest: '+ sheet.cell(row=r, column=3).value)
                                                                Making_differentials = (sheet.cell(row=r, column=4).value.split('|'))
                                                                for i in Making_differentials:
                                                                  List_of_Differentials.insert(END, i)

def Investigations():
        for i in Investigations_listbox.curselection():
                IxElements_Vs_Column_G.append(Investigations_listbox.get(i))
                Consult_Record.insert(END, Investigations_listbox.get(i) + '\n')

        
        for i in History:
                #wb = load_workbook('MDCheck.xlsx')
                for sheet in wb:
                        if sheet.title == i:
                                for r in range(1, 500):
                                        if sheet.cell(row=r, column=7).value != None and sheet.cell(row=r, column=8).value != None:
                                                for k in IxElements_Vs_Column_G:
                                                        if k == sheet.cell(row=r, column=3).value:
                                                                List_of_Differentials.insert(END, 'History suggest: '+ sheet.cell(row=r, column=8).value)
                                                                Making_differentials = (sheet.cell(row=r, column=9).value.split('|'))
                                                                for i in Making_differentials:
                                                                  List_of_Differentials.insert(END, i)


def Management():
        for i in Management_listbox.curselection():
                MxElements_Vs_Column_I.append(Management_listbox.get(i))
                Consult_Record.insert(END, Management_listbox.get(i) + '\n')

        
        for i in History:
                #wb = load_workbook('MDCheck.xlsx')
                for sheet in wb:
                        if sheet.title == i:
                                for r in range(1, 500):
                                        if sheet.cell(row=r, column=9).value != None and sheet.cell(row=r, column=10).value != None:
                                                for k in IxElements_Vs_Column_G:
                                                        if k == sheet.cell(row=r, column=9).value:
                                                                List_of_Differentials.insert(END, 'History suggest: '+ sheet.cell(row=r, column=9).value)
                                                                Making_differentials = (sheet.cell(row=r, column=10).value.split('|'))
                                                                for i in Making_differentials:
                                                                  List_of_Differentials.insert(END, i)


def legger():
  print( Consult_Record)
  print(Consult_Record.get(1.0, 'end'))

def history_differential():
  for i in List_of_Differentials.curselection():
                Consult_Record.insert(END, List_of_Differentials.get(i) + '\n')

def Clear():
  History_listbox.delete(0, 'end')
  Possible_differentials.delete(0,'end')
  Examination_listbox.delete(0, 'end')
  Investigations_listbox.delete(0, 'end')
  Management_listbox.delete(0, 'end')
  List_of_Differentials.delete(0, 'end')

def Update():
  Consult_Record.insert(END, update_detail.get()+'\n')
  

update = Button(root, text = 'Add detail', command= Update )
update.grid(column=4, row=2)
update_detail= Entry(root, width=30 )
update_detail.grid(column=3,row=2)

my_button = Button(root, text = '1log', command= logger)
my_button.grid(column=0, row=0)

OnRreview_button = Button(root, text = '2On review', command= On_review)
OnRreview_button.grid(column=1, row=0) 

OnRreview_of_DDX = Button(root, text = '3DDx Review', command= DDX)
OnRreview_of_DDX.grid(column=2, row=0)

OnExamination = Button(root, text = '4On Examination', command= On_PE)
OnExamination.grid(column=3, row=0)

onReview_of_Differentials= Button(root, text= 'Review of Differentials', command= history_differential)
onReview_of_Differentials.grid(column=3, row=3)

Ix = Button(root, text = '5Possible Ix for order', command= Investigations)
Ix.grid(column=4, row=0)

Mx = Button(root, text = '6For mx suggest', command= Management)
Mx.grid(column=3, row=4)

Record = Button(root, text = 'Record Print', command= legger )
Record.grid(column=4, row=3)

Clear = Button(root, text = 'Clear', command= Clear )
Clear.grid(column=4, row=4)

root.mainloop()

